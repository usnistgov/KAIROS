import os
import numpy as np
import pandas as pd
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
import copy

from kevs.produce_event_trees import get_ta2_predicted_events, \
    is_event_ins_or_pred


def format_graphg_worksheet(ws):
    # highlight in yellow predicted rows
    for row in ws.iter_rows():
        if row[3].value is True:
            for cell in row:
                cell.fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))

    # change column 2 width
    ws.column_dimensions['B'].width = 14
    # Start changing width from column C onwards
    column = 3
    while column <= ws.max_column:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = 33
        column += 1

    # Add texwrapping and top centering to all columns
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            alignment.vertical = 'top'
            cell.alignment = alignment

    # delete the True/False column
    ws.delete_cols(4, 1)


def produce_compare_event_str(ev_id: str, ta2_ce_instance, graphg_ceinstance) -> str:
    """
    Produce the event string for the ta2 event

    Args:
        ev_id:
        ta2_ce_instance:
        graphg_ceinstance:

    Returns:

    """
    dforig = ta2_ce_instance.ev_df
    g_nosuffix_df = graphg_ceinstance.ev_df
    # Only add Graph G column if not already in data frame
    if not ('graphg_ev_id' in dforig):
        if pd.notna(dforig['ev_provenance']).any():
            dforig = dforig. \
                merge(g_nosuffix_df.loc[:, ['graphg_ev_id']], how="left",
                      left_on="ev_provenance", right_on="graphg_ev_id")
        else:
            dforig['graphg_ev_id'] = np.nan

    if dforig.loc[dforig['ev_id'] == ev_id, 'ev_name'].tolist():
        ev_name_root = dforig.loc[dforig['ev_id'] == ev_id, 'ev_name'].iloc[0]
        ev_g_root = dforig.loc[dforig['ev_id'] == ev_id, 'graphg_ev_id'].iloc[0]
        if pd.isna(ev_g_root):
            ev_g_root = ""
        ev_pred_str = ""
        is_ins, is_pred = is_event_ins_or_pred(ev_id, ta2_ce_instance)
        if is_pred:
            ev_pred_str = "*predicted* "
        ev_conf_str = dforig.loc[dforig['ev_id'] == ev_id, 'ev_confidence_val'].iloc[0]
        ev_tree_name_root = "{}{} ({})".format(ev_pred_str, ev_name_root, ev_g_root)
        qlabel_str = dforig.loc[dforig['ev_id'] == ev_id, 'ev_ta2qlabel'].iloc[0]
        if pd.isna(qlabel_str) or qlabel_str == "":
            qlabel_str = dforig.loc[dforig['ev_id'] == ev_id, 'ev_qlabel'].iloc[0]
        desc_str = dforig.loc[dforig['ev_id'] == ev_id, 'ev_description'].iloc[0]
        ev_tree_name = "{}: description: {}, conf: {}".format(ev_tree_name_root,
                                                              desc_str, ev_conf_str)
    else:
        ev_tree_name = "**Event Name missing for event id {}**".format(ev_id)
    return ev_tree_name


def get_single_arg_string(arg_row):
    arg_name = arg_row['ent_name']
    arg_g_root = arg_row['graphg_ent_id']
    # RESIN cites the graph G entity differently from CMU and IBM
    if arg_row['ta2_team_name'] == "RESIN":
        arg_g_root = arg_row['arg_ta2provenance']
    if pd.isna(arg_g_root):
        arg_g_root = ""
    qlabel_str = arg_row['ent_TA2qlabel']
    if pd.isna(qlabel_str) or qlabel_str == "":
        qlabel_str = arg_row['ent_qlabel']
    arg_str = "{} ({})".format(arg_name, arg_g_root)
    return arg_str


def get_graph_g_entity(arg_ta2ent_str):
    try:
        graph_g_ent_id = arg_ta2ent_str.split('/')[-1]
    except IndexError:
        graph_g_ent_id = np.nan
    if arg_ta2ent_str == "kairos:NULL":
        graph_g_ent_id = np.nan
    return graph_g_ent_id


def produce_compare_arg_str(ev_id: str, ta2_ce_instance, graphg_ceinstance) -> str:
    """
    Produce the full information we wish to display for each event given its id in the event tree

    Args:
        ev_id:
        ta2_ce_instance:
        graphg_ceinstance:

    Returns:

    """
    dforig = ta2_ce_instance.ev_df
    g_nosuffix_df = graphg_ceinstance.ev_df
    # Only add graphg_ev_id if not already added
    if not ('graphg_ev_id' in dforig):
        if pd.notna(dforig['ev_provenance']).any():
            dforig = dforig. \
                merge(g_nosuffix_df.loc[:, ['graphg_ev_id']], how="left",
                      left_on="ev_provenance", right_on="graphg_ev_id")
        else:
            dforig['graphg_ev_id'] = np.nan

    if dforig.loc[dforig['ev_id'] == ev_id, 'ev_name'].tolist():
        dforigarg = ta2_ce_instance.arg_df
        # Since this is done repeatedly, please move above and pass this by reference
        dforigarg['graphg_ent_id'] = dforigarg['arg_ta2entity'].apply(get_graph_g_entity)

        arg_rows = dforigarg.loc[dforigarg['ev_id'] == ev_id, :]
        arg_rows = pd.merge(arg_rows,
                            ta2_ce_instance.ent_df.loc[:, ['ent_id', 'ent_name', 'ent_qnode',
                                                           'ent_qlabel', 'ent_TA2qnode',
                                                           'ent_TA2qlabel']], how="left",
                            left_on="arg_ta2entity", right_on="ent_id")
        arg_rows = arg_rows.loc[pd.notna(arg_rows["arg_ta2entity"]) &
                                (arg_rows["arg_ta2entity"] != "kairos:NULL"), :]
        if len(arg_rows) > 0:
            arg_rows['arg_row_str'] = arg_rows.apply(get_single_arg_string, axis=1)
            arg_str = ', '.join(arg_rows['arg_row_str'].tolist())
        else:
            arg_str = ""
        arg_str = "Ins. Args: " + arg_str
    else:
        arg_str = "**Event Name missing for event id {}**".format(ev_id)
    return arg_str


def get_graph_g_comparison(output_dir, annotation_collection, ta2_collection,
                           graph_g_collection,
                           task2_annotation_dir, match_dir):
    """
    Produce a graph G comparison between abridged and nosuffix events
    Args:
        output_dir:
        annotation_collection: task_1 annotation
        ta2_collection:
        graph_g_collection:
        task2_annotation_dir:
        match_dir:

    Returns:

    """
    ta2_team_list = ta2_collection.ta2dict.keys()
    ce_list = ['ce2013', 'ce2020', 'ce2024', 'ce2075', 'ce2079', 'ce2094',
               'ce2101', 'ce2102', 'ce2103', 'ce2104']
    # ce_list = ['ce2002']
    ta1_team_list = ['CMU', 'IBM', 'ISI', 'RESIN', 'SBU']
    # ta1_team_list = ['NISTTESTA']
    for ce in ce_list:
        abridgedce = ce + 'abridged'
        # Merge data frames for TA2 systems into one spreadsheet and write out

        # Get the Graph G Selected events file
        ep_selr_df = pd.read_excel(os.path.join(task2_annotation_dir, ce,
                                                ce + '_events_selected.xlsx'))
        ep_selr_df.rename(columns={'Unnamed: 0': "graphg_event_num",
                                   'eventprimitive_id': 'graphg_annotation_ev_id'},
                          errors="raise", inplace=True)

        # Get Graph G instances and annotation
        abridged_graphg_key = "GRAPHG|GRAPHG|{}|nist:Instances/00001/nistQuizlet9GraphG". \
            format(abridgedce)
        abridged_graphg_ceinstance = graph_g_collection.ta2dict['GRAPHG'].ta2dict[
            abridged_graphg_key]
        abridged_graphg_ceinstance.ev_df['graphg_ev_id'] = \
            abridged_graphg_ceinstance.ev_df['ev_id'].str.split('/', expand=True)[2]
        g_abridged_df = abridged_graphg_ceinstance.ev_df
        res_output_fname = "task2_graphgabridged_{}_comparison.xlsx".format(ce)
        res_output_path = os.path.join(output_dir, res_output_fname)
        ep_selr_df.to_excel(res_output_path, index=False,
                            sheet_name='Graph G')
        res_workbook = openpyxl.load_workbook(res_output_path)
        options = {}
        options['strings_to_formulas'] = False
        options['strings_to_urls'] = False
        writer = pd.ExcelWriter(res_output_path, engine='openpyxl', options=options)
        writer.book = res_workbook
        # Use the original index to sort at the end
        ep_selr_df.reset_index(drop=False, inplace=True)

        # Find the Annotation collection
        ann_ce = annotation_collection.annotation_dict['task1|{}'.format(ce)]
        ann_ev_df = ann_ce.ep_df.loc[:, ['eventprimitive_id', 'description']].merge(
            ep_selr_df.loc[:, ['graphg_event_num', 'graphg_annotation_ev_id']],
            left_on='eventprimitive_id', right_on="graphg_annotation_ev_id",
            how="left")
        ann_ev_df.sort_values(by=['graphg_event_num'], ascending=True, inplace=True)
        ann_ev_df.drop(columns=['graphg_annotation_ev_id'], inplace=True)
        for ta2_team in ta2_team_list:
            ta2_instance = ta2_collection.ta2dict[ta2_team]

            # Reset the data frame for each TA2 system:
            final_ce_df = ann_ev_df.copy(deep=True)

            # Make a data frame per (ta2, ce), with each row being a TA1 team
            for ta1_team in ta1_team_list:
                ta2_abridgedce_instance_list = [(key, value) for
                                                key, value in ta2_instance.ta2dict.items() if
                                                ((ta1_team in key.split('|')[0]) and
                                                 (ta2_team in key.split('|')[1]) and
                                                 (abridgedce in key.split('|')[2]))]
                for (key, value) in ta2_abridgedce_instance_list:
                    abridged_ta2_ceinstance = value
                    final_df_col_name_base = abridged_ta2_ceinstance.ce_instance_file_name_base
                    abridged_df = abridged_ta2_ceinstance.ev_df
                    if pd.notna(abridged_df['ev_provenance']).any():
                        abridged_df = abridged_ta2_ceinstance.ev_df. \
                            merge(g_abridged_df.loc[:, ['graphg_ev_id']], how="left",
                                  left_on="ev_provenance", right_on="graphg_ev_id")
                    else:
                        abridged_df['graphg_ev_id'] = np.nan
                    if pd.notna(abridged_ta2_ceinstance.ev_df['ev_provenance']).any():
                        abridged_ta2_ceinstance.ev_df = abridged_ta2_ceinstance.ev_df. \
                            merge(g_abridged_df.loc[:, ['graphg_ev_id']], how="left",
                                  left_on="ev_provenance", right_on="graphg_ev_id")
                    else:
                        abridged_df['graphg_ev_id'] = np.nan
                    abridged_pred_events = get_ta2_predicted_events(abridged_df)
                    abridged_df['abridged_predicted_event'] = False
                    if len(abridged_df.loc[abridged_df['ev_id'].isin(abridged_pred_events['ev_id']),
                                           'abridged_predicted_event']) > 0:
                        abridged_df.loc[abridged_df['ev_id'].isin(abridged_pred_events['ev_id']),
                                        'abridged_predicted_event'] = True

                    # Now produce event string
                    abridged_df = abridged_df.loc[abridged_df['ev_ta1ref'] != "kairos:NULL", :]

                    # Get the matching
                    match_fpath = \
                        os.path.join(match_dir, ta2_team,
                                     "{}_matched_events.csv".format(
                                         abridged_ta2_ceinstance.ce_instance_file_name_base))
                    match_df = pd.read_csv(match_fpath)

                    abridged_df["abridged_event_str"] = abridged_df['ev_id'].apply(
                        produce_compare_event_str,
                        args=(abridged_ta2_ceinstance,
                              abridged_graphg_ceinstance,))
                    abridged_df["abridged_arg_str"] = abridged_df['ev_id'].apply(
                        produce_compare_arg_str,
                        args=(abridged_ta2_ceinstance,
                              abridged_graphg_ceinstance,))
                    abridged_df['abridged_evarg_str'] = \
                        abridged_df['abridged_event_str'] + '; ' + \
                        abridged_df['abridged_arg_str']

                    sim_metric_list = ['class', 'complex', 'jc', 'text', 'topsim', 'transe']
                    for metric in sim_metric_list:
                        final_df_col_name = final_df_col_name_base + "_" + metric
                        mm_df = match_df.loc[pd.notna(match_df[metric]),
                                             ['ta2_ev_id', 'ann_ev_id', metric]]
                        mm_df = mm_df.merge(abridged_df.loc[:, ['ev_id', 'abridged_evarg_str']],
                                            left_on="ta2_ev_id", right_on='ev_id',
                                            how='left')
                        mm_df.drop(columns=['ev_id'], inplace=True)
                        res_df = ann_ev_df.merge(mm_df, how="left",
                                                 left_on="eventprimitive_id",
                                                 right_on="ann_ev_id")
                        res_df['sim_str_front'] = '[sim: '
                        res_df['sim_str_back'] = '] '

                        res_df['abridged_evarg_sim_str'] = res_df['sim_str_front'] + \
                            res_df[metric].astype(str) + \
                            res_df['sim_str_back'] + \
                            res_df['abridged_evarg_str']
                        res_df = res_df.loc[:, ['eventprimitive_id', 'abridged_evarg_sim_str']]
                        # Now add to the final df
                        final_ce_df = final_ce_df.merge(res_df, how="left", on="eventprimitive_id",)
                        final_ce_df.rename(columns={'abridged_evarg_sim_str': final_df_col_name},
                                           errors="raise", inplace=True)
                        final_ce_df.sort_values(by=['graphg_event_num', 'eventprimitive_id'],
                                                ascending=True, inplace=True)

                # We add rows in the same sheet for ta1
            # After all of the instances give us a row, write a sheet for that TA2 team
            final_ce_df.to_excel(writer, index=False,
                                 sheet_name='{}'.format(ta2_team))
        # Format the columns and sells
        for sheet in ['IBM', 'CMU', 'RESIN']:
            format_graphg_worksheet(res_workbook[sheet])
        # Write the output to the file and close it
        writer.save()
        writer.close()
